from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook, load_workbook

from zk_zone_agent.device_users import MAX_NAME_BYTES


CNIC_SUFFIX_RE = re.compile(r"^(?P<name>.*?)(?P<shift>-S)?-(?P<cnic>\d{13})$")
EXPECTED_HEADERS = ("ID", "name", "CNIC")


@dataclass(frozen=True)
class ExportUserRow:
    user_id: str
    name: str
    cnic: str


@dataclass(frozen=True)
class ParsedBulkUserRow:
    row_number: int
    user_id: str
    sheet_name: str
    cnic: str | None
    status: str
    message: str | None
    expected_name: str | None


@dataclass(frozen=True)
class MachineIdentity:
    employee_name: str
    cnic: str
    raw_punch: bool


def split_machine_identity(name: str | None) -> MachineIdentity:
    normalized = normalize_name(name or "")
    match = CNIC_SUFFIX_RE.match(normalized)
    if not match:
        return MachineIdentity(normalized, "", False)
    return MachineIdentity(
        employee_name=normalize_name(match.group("name")),
        cnic=match.group("cnic"),
        raw_punch=bool(match.group("shift")),
    )


def split_machine_name_cnic(name: str | None) -> tuple[str, str]:
    identity = split_machine_identity(name)
    return identity.employee_name, identity.cnic


def normalize_name(value: str) -> str:
    return " ".join(str(value or "").strip().split())


def normalize_cnic(value: object) -> str:
    text = _cell_text(value)
    digits = "".join(char for char in text if char.isdigit())
    if not digits:
        return ""
    if len(digits) != 13:
        raise ValueError("CNIC must contain exactly 13 digits.")
    return digits


def build_machine_name(name: str, cnic: str) -> str:
    name = normalize_name(name)
    if not name:
        raise ValueError("Name is required when CNIC is provided.")
    suffix = f"-{cnic}"
    prefix_limit = MAX_NAME_BYTES - len(suffix.encode("utf-8"))
    if prefix_limit < 1:
        raise ValueError("CNIC suffix does not fit within the device name limit.")
    prefix = _truncate_utf8(name, prefix_limit).strip()
    if not prefix:
        raise ValueError("Name cannot fit before CNIC within the device name limit.")
    final_name = f"{prefix}{suffix}"
    if len(final_name.encode("utf-8")) > MAX_NAME_BYTES:
        raise ValueError(f"Name must fit within {MAX_NAME_BYTES} UTF-8 bytes for the device.")
    return final_name


def export_users_xlsx(rows: Iterable[ExportUserRow]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Device Users"
    sheet.append(EXPECTED_HEADERS)
    for row in rows:
        sheet.append([row.user_id, row.name, row.cnic])
    for column, width in {"A": 18, "B": 32, "C": 20}.items():
        sheet.column_dimensions[column].width = width
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def parse_bulk_update_xlsx(content: bytes) -> list[ParsedBulkUserRow]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Upload a valid .xlsx file.") from exc
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The uploaded workbook is empty.")
    headers = tuple(_cell_text(value) for value in rows[0][:3])
    if headers != EXPECTED_HEADERS:
        raise ValueError("The first row must contain exactly: ID, name, CNIC.")

    parsed: list[ParsedBulkUserRow] = []
    seen_ids: set[str] = set()
    for index, values in enumerate(rows[1:], start=2):
        user_id = _cell_text(values[0] if len(values) > 0 else "")
        sheet_name = normalize_name(_cell_text(values[1] if len(values) > 1 else ""))
        raw_cnic = values[2] if len(values) > 2 else ""
        if not user_id and not sheet_name and not _cell_text(raw_cnic):
            continue
        if not user_id:
            parsed.append(_failed_row(index, user_id, sheet_name, "ID is required."))
            continue
        if user_id in seen_ids:
            parsed.append(_failed_row(index, user_id, sheet_name, f"Duplicate ID {user_id} in upload."))
            continue
        seen_ids.add(user_id)
        try:
            cnic = normalize_cnic(raw_cnic)
        except ValueError as exc:
            parsed.append(_failed_row(index, user_id, sheet_name, str(exc)))
            continue
        if not cnic:
            parsed.append(
                ParsedBulkUserRow(
                    row_number=index,
                    user_id=user_id,
                    sheet_name=sheet_name,
                    cnic=None,
                    status="SKIPPED",
                    message="CNIC is blank; row skipped.",
                    expected_name=None,
                )
            )
            continue
        try:
            expected_name = build_machine_name(sheet_name, cnic)
        except ValueError as exc:
            parsed.append(_failed_row(index, user_id, sheet_name, str(exc), cnic=cnic))
            continue
        parsed.append(
            ParsedBulkUserRow(
                row_number=index,
                user_id=user_id,
                sheet_name=sheet_name,
                cnic=cnic,
                status="PENDING",
                message=None,
                expected_name=expected_name,
            )
        )
    if not parsed:
        raise ValueError("The uploaded workbook does not contain any user rows.")
    return parsed


def _failed_row(
    row_number: int,
    user_id: str,
    sheet_name: str,
    message: str,
    *,
    cnic: str | None = None,
) -> ParsedBulkUserRow:
    return ParsedBulkUserRow(
        row_number=row_number,
        user_id=user_id,
        sheet_name=sheet_name,
        cnic=cnic,
        status="FAILED",
        message=message,
        expected_name=None,
    )


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _truncate_utf8(value: str, byte_limit: int) -> str:
    output: list[str] = []
    used = 0
    for char in value:
        char_bytes = len(char.encode("utf-8"))
        if used + char_bytes > byte_limit:
            break
        output.append(char)
        used += char_bytes
    return "".join(output)
